#!/usr/bin/env python3

"""
导出全部题材股票明细到Excel
功能：遍历所有题材，获取每个题材的股票明细，导出到Excel文件
"""

import os
import json
import sys
import time
from typing import Any
from pathlib import Path

# 设置标准输出编码为utf-8，防止Windows下打印特殊字符报错
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

import pandas as pd
import requests
from get_block_stocks import fetch_theme_list
from openpyxl import load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Load environment variables from .env file if it exists
env_file = Path(__file__).parent / ".env"
if env_file.exists():
    with open(env_file) as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                key, value = line.split("=", 1)
                os.environ[key.strip()] = value.strip()

# API配置 - 从环境变量读取，如果不存在则使用默认值
USER_ID = os.getenv("LONGHUVIP_USER_ID", "397605")
DEVICE_ID = os.getenv("LONGHUVIP_DEVICE_ID", "548d826f-a2a7-301a-b148-920f31f15331")
TOKEN = os.getenv("LONGHUVIP_TOKEN", "df9cadb87bbba7d04e9fcbaa2aa229b3")

API_BASE = (
    f"https://applhb.longhuvip.com/w1/api/index.php?"
    f"a=InfoGet&apiv=w43&c=Theme&PhoneOSNew=1&"
    f"UserID={USER_ID}&DeviceID={DEVICE_ID}&"
    f"VerSion=5.22.0.2&Token={TOKEN}&ID="
)


def get_theme_detail(theme_id: str, debug: bool = False) -> dict[str, Any]:
    """获取单个题材的详细信息"""
    url = API_BASE + theme_id
    try:
        if debug:
            print(f"\n    [DEBUG] 请求URL: {url}")
            print(
                f"    [DEBUG] 题材ID: {theme_id!r} (长度: {len(theme_id)}, 类型: {type(theme_id)})"
            )
            # 打印ID的十六进制表示（如果是字节）
            if isinstance(theme_id, bytes):
                print(f"    [DEBUG] 题材ID(hex): {theme_id.hex()}")

        response = requests.get(url, timeout=30)

        if debug:
            print(f"    [DEBUG] HTTP状态码: {response.status_code}")
            print(f"    [DEBUG] 响应头: {dict(response.headers)}")

        response.raise_for_status()
        data = response.json()

        if debug:
            print(f"    [DEBUG] API响应: {json.dumps(data, ensure_ascii=False, indent=2)[:500]}...")

        if data.get("errcode") != "0":
            errcode = data.get("errcode")
            errmsg = data.get("errmsg", "")
            raise ValueError(f"API返回错误: {errcode}" + (f" - {errmsg}" if errmsg else ""))

        return data
    except Exception as e:
        if debug:
            print(f"    [DEBUG] 异常类型: {type(e).__name__}")
            print(f"    [DEBUG] 异常详情: {str(e)}")
        raise Exception(f"获取题材 {theme_id} 详情失败: {str(e)}")


def parse_stock_data(data: dict[str, Any], theme_name: str) -> list[dict[str, str]]:
    """
    解析股票数据，返回列表
    返回格式: [{theme_name, level1, level2, stock_id, stock_name, reason}, ...]
    """
    stocks_data = []

    # 情况1：Table为空但根级有Stocks
    if (not data.get("Table") or len(data.get("Table", [])) == 0) and data.get("Stocks"):
        for stock in data.get("Stocks", []):
            stocks_data.append(
                {
                    "题材名称": theme_name,
                    "第一层级": "",
                    "第二层级": "",
                    "股票代码": stock.get("StockID", ""),
                    "股票名称": stock.get("prod_name", ""),
                    "关联原因": stock.get("Reason", ""),
                }
            )

    # 情况2：有Table层级结构
    elif data.get("Table") and len(data.get("Table", [])) > 0:
        for table_item in data.get("Table", []):
            if table_item.get("Level1"):
                level1 = table_item.get("Level1")
                level1_name = level1.get("Name", "")

                # Level1自身的Stocks
                if level1.get("Stocks"):
                    for stock in level1.get("Stocks", []):
                        stocks_data.append(
                            {
                                "题材名称": theme_name,
                                "第一层级": level1_name,
                                "第二层级": "",
                                "股票代码": stock.get("StockID", ""),
                                "股票名称": stock.get("prod_name", ""),
                                "关联原因": stock.get("Reason", ""),
                            }
                        )

                # Level2子层级
                if table_item.get("Level2"):
                    for level2 in table_item.get("Level2", []):
                        level2_name = level2.get("Name", "")
                        if level2.get("Stocks"):
                            for stock in level2.get("Stocks", []):
                                stocks_data.append(
                                    {
                                        "题材名称": theme_name,
                                        "第一层级": level1_name,
                                        "第二层级": level2_name,
                                        "股票代码": stock.get("StockID", ""),
                                        "股票名称": stock.get("prod_name", ""),
                                        "关联原因": stock.get("Reason", ""),
                                    }
                                )

    return stocks_data


def sanitize_text(value: Any) -> str:
    """
    清理文本，去除 Excel 不允许的控制字符，并确保返回字符串
    """
    if value is None:
        return ""
    text = str(value)
    # openpyxl 会拒绝包含以下控制字符的数据，提前移除以避免写入失败
    return ILLEGAL_CHARACTERS_RE.sub("", text)


def sanitize_row(row: dict[str, Any]) -> dict[str, str]:
    """清理单行数据中的所有字段"""
    return {k: sanitize_text(v) for k, v in row.items()}


def format_progress(current: int, total: int, theme_name: str) -> str:
    """格式化进度信息"""
    percentage = (current / total * 100) if total > 0 else 0
    return f"[{current}/{total}] ({percentage:.1f}%) {theme_name}"


def export_all_themes_to_excel(output_file: str = "全部题材股票明细.xlsx"):
    """
    导出全部题材股票明细到Excel

    Args:
        output_file: 输出文件名，默认为"全部题材股票明细.xlsx"
    """
    print("=" * 80)
    print("开始导出全部题材股票明细")
    print("=" * 80)

    # 1. 获取题材列表
    print("\n[步骤 1/4] 正在获取题材列表...")
    try:
        themes = fetch_theme_list()
        if not themes:
            print("[ERROR] 获取题材列表失败或列表为空")
            return

        print(f"[OK] 成功获取 {len(themes)} 个题材")

        # 检查并过滤异常数据
        valid_themes = []
        invalid_count = 0
        for theme in themes:
            theme_id = theme.get("id", "")
            theme_name = theme.get("name", "")

            # 验证ID是否为纯数字
            if not theme_id or not str(theme_id).strip().isdigit():
                invalid_count += 1
                continue

            # 验证名称
            if not theme_name or len(theme_name) == 0 or len(theme_name) > 100:
                invalid_count += 1
                continue

            valid_themes.append(theme)

        if invalid_count > 0:
            print(f"[WARN] 过滤掉 {invalid_count} 个异常题材")

        themes = valid_themes
        if not themes:
            print("[ERROR] 没有有效的题材数据")
            return

    except Exception as e:
        print(f"[ERROR] 获取题材列表失败: {str(e)}")
        return

    # 2. 遍历所有题材，获取详细信息
    print("\n[步骤 2/4] 正在获取各题材详细信息...")
    all_rows = []
    success_count = 0
    fail_count = 0

    total = len(themes)
    for i, theme in enumerate(themes, 1):
        theme_id = theme.get("id", "")
        theme_name = theme.get("name", "未知题材")
        theme_code = theme.get("code", "")

        # 打印进度
        progress = format_progress(i, total, theme_name)
        print(f"  {progress}", end="", flush=True)

        try:
            # 获取题材详情
            detail_data = get_theme_detail(theme_id, debug=False)

            # 解析股票数据
            stocks_data = parse_stock_data(detail_data, theme_name)
            # 清理非法字符，避免写入 Excel 失败
            stocks_data = [sanitize_row(row) for row in stocks_data]

            if stocks_data:
                all_rows.extend(stocks_data)
                print(f" [OK] ({len(stocks_data)} 只股票)")
                success_count += 1
            else:
                # 没有股票数据也记录一行
                all_rows.append(
                    sanitize_row(
                        {
                            "题材名称": theme_name,
                            "第一层级": "",
                            "第二层级": "",
                            "股票代码": "",
                            "股票名称": "",
                            "关联原因": "没有可用的股票数据",
                        }
                    )
                )
                print(" [WARN] (无股票数据)")
                success_count += 1

            # 避免请求过快，添加短暂延迟
            if i < total:
                time.sleep(0.2)

        except Exception as e:
            error_msg = str(e)
            # 简化错误信息显示
            if len(error_msg) > 100:
                error_msg = error_msg[:100] + "..."
            print(f" [ERROR] 失败: {error_msg}")
            fail_count += 1

            # 记录失败信息
            all_rows.append(
                sanitize_row(
                    {
                        "题材名称": theme_name,
                        "第一层级": "",
                        "第二层级": "",
                        "股票代码": "",
                        "股票名称": "",
                        "关联原因": f"获取详细信息失败: {str(e)}",
                    }
                )
            )

    print(
        f"\n[OK] 数据获取完成: 成功 {success_count} 个，失败 {fail_count} 个，共 {len(all_rows)} 条记录"
    )

    # 3. 创建DataFrame并导出到Excel
    print("\n[步骤 3/4] 正在生成Excel文件...")
    try:
        if not all_rows:
            print("[ERROR] 没有数据可导出")
            return

        df = pd.DataFrame(all_rows)

        # 确保列的顺序
        columns = ["题材名称", "第一层级", "第二层级", "股票代码", "股票名称", "关联原因"]
        df = df[columns]

        # 导出到Excel
        df.to_excel(output_file, index=False, engine="openpyxl")
        print(f"[OK] Excel文件已生成: {output_file}")

    except Exception as e:
        print(f"[ERROR] 生成Excel文件失败: {str(e)}")
        return

    # 4. 美化Excel格式
    print("\n[步骤 4/4] 正在美化Excel格式...")
    try:
        wb = load_workbook(output_file)
        ws = wb.active

        # 设置表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # 设置列宽
        column_widths = {
            "题材名称": 25,
            "第一层级": 20,
            "第二层级": 20,
            "股票代码": 12,
            "股票名称": 20,
            "关联原因": 40,
        }

        for col_idx, col_name in enumerate(columns, 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = column_widths.get(col_name, 15)

        # 设置数据行对齐
        data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = data_alignment

        # 冻结首行
        ws.freeze_panes = "A2"

        # 保存
        wb.save(output_file)
        print("[OK] Excel格式美化完成")

    except Exception as e:
        print(f"[WARN] 美化Excel格式失败（文件已生成）: {str(e)}")

    # 5. 完成
    print("\n" + "=" * 80)
    print("导出完成！")
    print("[STATS] 统计信息:")
    print(f"   - 题材总数: {total}")
    print(f"   - 成功导出: {success_count}")
    print(f"   - 失败数量: {fail_count}")
    print(f"   - 股票记录: {len(all_rows)}")
    print(f"   - 输出文件: {output_file}")
    print("=" * 80)


if __name__ == "__main__":
    import sys

    # 支持命令行参数指定输出文件名
    output_file = sys.argv[1] if len(sys.argv) > 1 else "全部题材股票明细.xlsx"

    try:
        export_all_themes_to_excel(output_file)
    except KeyboardInterrupt:
        print("\n\n[WARN] 用户中断操作")
        sys.exit(1)
    except Exception as e:
        print(f"\n[ERROR] 发生错误: {str(e)}")
        import traceback

        traceback.print_exc()
        sys.exit(1)
